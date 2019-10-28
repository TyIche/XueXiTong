#LittlePrincess
i=1
path=''
emm=['1','0','2','3']
import openpyxl
workbookin=openpyxl.Workbook()
worksheetin=workbookin.active
import os 


#global path
#global i
def makein( tot , id ,pts,ws):
    #global path
    #print(id)
    for row in ws.rows:
        if(str(list(row)[0].value)==id):
            hang=row[0].row
            worksheetin.cell(row=hang,column=i+2,value=pts)
            #print('hao',pts,hang)
            break
def work0( pathh,ws ):
    global i
    workbook = openpyxl.load_workbook(pathh)
    worksheet = workbook.worksheets[0]
    for row in worksheet.rows:
        #print(int(str(list(row)[6].value)))
        if(str(list(row)[6].value) in emm):
            makein(i,str(list(row)[0].value),1,ws)
        else:
            makein(i,str(list(row)[0].value),0,ws)
    i+=1
def work1(pathh,ws):
    global i
    workbook = openpyxl.load_workbook(pathh)
    worksheet = workbook.worksheets[0]
    for row in worksheet.rows:
        #print(row[0].row)
        makein(i,str(list(row)[0].value),str(list(row)[6].value),ws)
    i+=1
def work2(pathh,ws):
    global i
    workbook = openpyxl.load_workbook(pathh)
    worksheet = workbook.worksheets[0]
    for row in worksheet.rows:
        #print(str(list(row)[5].value))
        if(str(list(row)[5].value)=='已签'):
            makein(i,str(list(row)[1].value),1,ws)
        else:
            makein(i,str(list(row)[1].value),0,ws)
    i+=1
def main():
    global path
    path=input("输入绝对路径：")
    #path+=r'/'
    #print('wdnmd')
    i=1
    listt = os.listdir(path)
    #print(listt)
    flag=0
    for j,filename in enumerate(listt):
        #print('wdnmd')
        if(flag==0):
            flag=1
            eg=path+r'/'+filename
            EG=openpyxl.load_workbook(eg)
            ws=EG.worksheets[0]
        #print(filename)
        if(filename[0]=='~'):
            continue
        if(filename[-7]=='签'):
            work0(path+r'/'+filename,ws)
            print(filename,' ',1)
        else:
            if(filename[0]=='签'):
                work2(path+r'/'+filename,ws)
                print(filename,' ',2)
            else:
                work1(path+r'/'+filename,ws)
                print(filename,' ',3)
    workbookin.save('out.xlsx')
main()
